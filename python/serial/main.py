# -*- coding: utf-8 -*-
import PySimpleGUI as sg


import threading
import math
import serial
import serial.tools.list_ports
import time
import re

import xlwt
#import openpyxl
import datetime

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Border, Side, Alignment

global ser

def hex_2_short(hexstr):
    width = 16
    intnum=int(hexstr, 16)
    if intnum > 2 ** (width-1)- 1:
        intnum = 2 ** width - intnum
        intnum=0-intnum
    return intnum

def serialopen(portx):
    global serial_isopen
    bps=230400
    timex=5
    #串口执行到这已经打开 再用open命令会报错
    ser = serial.Serial(portx, int(bps), timeout=1, parity=serial.PARITY_NONE,stopbits=1)
    ser.flushInput() # 清空缓冲区
    if (ser.isOpen()):
        print("open success")
        # 向端口些数据 字符串必须译码
        #ser.write("hello".encode())
        serial_isopen = 1
    else:
        print("open failed")
        ser.close()#关闭端口
    return ser

flag  = 1
num = 0
def run():
    print(threading.currentThread().getName() + '\n')
    global flag
    global ser
    #global worksheet
    #global style
    #global num
    global ws
    ser.flushInput() # 清空缓冲区
    flag = 1
    while True:
        #while (True):
        #line = ser.readline()
        retstr = ser.read(16).hex()
        #print(retstr)
        #print(type(retstr))
        speed = retstr[20:32];
        speedx =speed[2:4] + speed[0:2];
        speedy =speed[6:8] + speed[4:6];
        speedz =speed[10:12] + speed[8:10];

        speedx = hex_2_short(speedx)
        speedy = hex_2_short(speedy)
        speedz = hex_2_short(speedz)
        speedx = (speedx/32768)*16
        speedy = (speedy/32768)*16
        speedz = (speedz/32768)*16
        print ("x:",speedx)
        print ("y:",speedy)
        print ("z:",speedz)
        speed =  math.sqrt(speedx*speedx + speedy*speedy +speedz*speedz)
        #print(type(speed))
        print( time.strftime("%Y-%m-%d %X",time.localtime())," --- speed --> ", speed) # 打印一下
        '''
        worksheet.write(num, 0, datetime.datetime.now(), style)
        worksheet.write(num, 1, speed)
        worksheet.write(num, 2, speedx)
        worksheet.write(num, 3, speedy)
        worksheet.write(num, 4, speedz)
        num  = num +  1
        '''
        ws.append([datetime.datetime.now(), speed, speedx, speedy, speedz])
        # 加入延时，不跑满带宽
        time.sleep(0.1)
        if flag != 1 :
            break

def serialclose():
    global serial_isopen
    ser.close()#关闭端口
    serial_isopen = 0
    print("close success")
filepath = ''
#com=('COM1','COM2','COM3','COM4','COM5','COM6','COM7','COM8','COM9','COM10')
com = list(serial.tools.list_ports.comports())
# Create some widgets
text = sg.Text("输入设备COM号：",size=(15, 1))
text_entry = sg.InputCombo(com,size=(20, 1),key='-LIST-')
#text_entry = sg.InputText()
ok_btn = sg.Button('open')
close_btn = sg.Button('close')
start_btn = sg.Button('start')
stop_btn = sg.Button('stop')

exit_btn = sg.Button('Exit')
#file_entry = sg.FileSaveAs(
#        key='fig_save',
#        #file_types=(('Excel', '.xls')),  # TODO: better names
#        #file_types=(("Mp3 Files", "*.mp3")),
#        )
menu_def = [['File', ['Open', 'Save', 'Exit',]],
            ['log', ['on::_LOG_ON_', 'off::_LOG_OFF_','set::_LOG_SET_']],
            ['Help', 'About...'],]

#layout = [[sg.Menu(menu_def)],
#          [sg.Text('', size=(60, 1))],
#          [text, text_entry, ok_btn, close_btn],
#          [start_btn,stop_btn],
#          [exit_btn]]
layout = [[text, text_entry, ok_btn, close_btn],
          [sg.Text('日志保存位置:', size=(15, 1)), sg.InputText(), sg.FolderBrowse(key='-FILE_BROWSE-')],
          [start_btn,stop_btn],
          [sg.Output(size=(65,5),key='-TEXT_OUT-')],
          [exit_btn]]

# Create the Window
window = sg.Window('Hello Speed', layout)
'''
excel 文件相关
'''
# 设置excel文件相关
'''
workbook = xlwt.Workbook()
worksheet = workbook.add_sheet('My Sheet')
#worksheet.column_dimensions['A'].width = 23
style = xlwt.XFStyle()
style.num_format_str = 'M/D/YY h:mm:ss' # Other options: D-MMM-YY, D-MMM, MMM-YY, h:mm, h:mm:ss, h:mm, h:mm:ss, M/D/YY h:mm, mm:ss, [h]:mm:ss, mm:ss.0
worksheet.write(num, 0, "时间")
worksheet.write(num, 1, "加速度(g/s)")
worksheet.write(num, 2, "X轴加速度(g/s)")
worksheet.write(num, 3, "Y轴加速度(g/s)")
worksheet.write(num, 4, "Z轴加速度(g/s)")
num = num + 1;
'''

wb=Workbook()  #创建一个工作表
ws=wb.active   #ws操作sheet页
#ws = wb.create_sheet()
ws.append(['时间', '加速度(g/s)', 'X轴加速度(g/s)','Y轴加速度(g/s)','Z轴加速度(g/s)'])
ws.column_dimensions['A'].width = 23
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15



filename = ''
serial_isopen = 0
serial_isstart = 0
# Create the event loop
while True:
    #value = window.ReadNonBlocking()
    event, values = window.read()
    if event in (None, 'Exit'):
        # User closed the Window or hit the Cancel button
        break
    if event in (None, 'open'):
        portlist = values['-LIST-']
        comport = "".join(portlist) #转换为字符串
        comport = re.findall(r"[(](.*?)[)]",comport)
        ser = serialopen(comport[0])

    if event in (None, 'close'):
        serialclose()
    if event in (None, 'start'):
        if serial_isopen == 0 :
            sg.popup('Error', '串口未打开', text)
            continue
        if serial_isstart == 1 :
            sg.popup('Error', '任务已开始', text)
            continue
        #seri.serialread(ser)
        #注意这,开始咯,指明具体的方法和方法需要的参数
        #my_thread = threading.Thread(target=run, args=(ser,))
        my_thread = threading.Thread(target=run)
        #run()
        #filepath = values['-FILE_BROWSE-']
        #filename = time.strftime("%Y%m%d%H%M%S",time.localtime())
        #filename = filepath + '/' + filename + ".xlsx"

        #一定不要忘记
        my_thread.start()
        serial_isstart = 1
        window['-TEXT_OUT-'].update("start ... ")
        print("start ... ")

    if event in (None, 'stop'):
        filepath = values['-FILE_BROWSE-']
        flag = False
        if serial_isstart == 0 :
            sg.popup('Error', '任务未开始', text)
            continue
        print("stop ... ")
        filename = time.strftime("%Y%m%d%H%M%S",time.localtime())
        filename = filepath + '/' + filename + ".xlsx"
        wb.save(filename)
        print("文件保存在：", filename)
        my_thread.join()
        #workbook.save(filename)
        serial_isstart = 0
    if event in (None, 'on::_LOG_SET_'):
        filepath = sg.popup_get_folder('Please enter a folder name')
        print (filepath)
        #sg.popup('Results', 'The value returned from popup_get_folder', text)

    #print(f'Event: {event}')
    #print(str(values))

window.close()